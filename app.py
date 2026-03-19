from flask import Flask, render_template, request, send_file
import openpyxl
import os
import zipfile
import re

app = Flask(__name__)

TEMPLATE_FOLDER = "excel_templates"
OUTPUT_FOLDER = "output"

os.makedirs(OUTPUT_FOLDER, exist_ok=True)


# =========================
# 扫描Excel模板变量
# =========================
def get_variables():

    variables = set()

    for file in os.listdir(TEMPLATE_FOLDER):

        if file.endswith(".xlsx"):

            path = os.path.join(TEMPLATE_FOLDER, file)

            wb = openpyxl.load_workbook(path)

            sheet = wb.active

            for row in sheet.iter_rows():

                for cell in row:

                    if cell.value:

                        found = re.findall(r"\{\{(.*?)\}\}", str(cell.value))

                        for f in found:

                            variables.add(f)

    return list(variables)


# =========================
# 发票信息自动填充
# =========================
invoice_map = {

    "海田控股": {
        "name": "宁波海田控股集团有限公司",
        "tax": "330204567041225",
        "address": "会展路181号宁波国际贸易展览中心87349429",
        "bank": "交通银行宁波市分行332006271018010103062"
    },

    "大洲进出口": {
        "name": "宁波大洲进出口有限公司",
        "tax": "91330203768515570K",
        "address": "宁波市海曙区粮丰街吴黄1幢1-31室0574-87369425",
        "bank": "宁波银行月湖支行22020122000019594"
    }

}


# =========================
# 首页
# =========================
@app.route("/")
def index():

    variables = get_variables()

    return render_template("index.html", variables=variables)


# =========================
# 生成Excel文件
# =========================
@app.route("/generate", methods=["POST"])
def generate():

    data = request.form.to_dict()

    # 允许空字段
    data.setdefault("name","")

    # 自动填充发票信息
    title = data.get("title")

    if title in invoice_map:

        data["name"] = invoice_map[title]["name"]
        data["tax"] = invoice_map[title]["tax"]
        data["address"] = invoice_map[title]["address"]
        data["bank"] = invoice_map[title]["bank"]

    generated_files = []

    for template in os.listdir(TEMPLATE_FOLDER):

        if not template.endswith(".xlsx"):
            continue

        template_path = os.path.join(TEMPLATE_FOLDER, template)

        wb = openpyxl.load_workbook(template_path)

        sheet = wb.active

        for row in sheet.iter_rows():

            for cell in row:

                if cell.value:

                    for key, value in data.items():

                        cell.value = str(cell.value).replace("{{"+key+"}}", value)

        output_path = os.path.join(OUTPUT_FOLDER, template)

        wb.save(output_path)

        generated_files.append(output_path)


    # ZIP文件名 = Contract_No.
    contract_no = data.get("Contract_No.", "documents")

    zip_path = os.path.join(OUTPUT_FOLDER, f"{ ontract_no}.zip")

    with zipfile.ZipFile(zip_path, "w") as zipf:

        for file in generated_files:

            zipf.write(file, os.path.basename(file))

    return send_file(zip_path, as_attachment=True)


# =========================
# Render启动
# =========================
if __name__ == "__main__":

    port = int(os.environ.get("PORT", 10000))

    app.run(host="0.0.0.0", port=port)
